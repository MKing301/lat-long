import sys
import os
import openpyxl
import requests
import logging

from rich.traceback import install
from rich.console import Console
from dotenv import load_dotenv
from rich.progress import track


load_dotenv(os.getenv('ENV_PATH')) # Load environment variables
install() # Used for debugging
console = Console()

logging.basicConfig(
    filename='/home/non-prod-server/scripts/lat-long/logs/latlong.log',
    format='%(asctime)s %(message)s',
    datefmt='%m/%d/%Y %I:%M:%S %p',
    encoding='utf-8',
    level=logging.DEBUG

)

API_KEY = os.getenv('GOOGLE_MAP_API_KEY')


def get_lat_long(addr, key):
    try:
        response = requests.get(f'https://maps.googleapis.com/maps/api/geocode/json?address={addr}&key={key}')
        if response.status_code == 200:
            data = response.json()
            if data['status'] == 'OK':
                latitude = data['results'][0]['geometry']['location']['lat']
                longitude = data['results'][0]['geometry']['location']['lng']
                formatted_addr = data['results'][0]['formatted_address']
                return (latitude, longitude, formatted_addr)
            else:
                logging.error(f'{data["status"]}: {data["error_message"]}')
                sys.exit()
        return (None, None, None)
    except Exception as e:
        logging.error(e)



logging.info('Starting script ...')

# Load workbook
try:
    wb = openpyxl.load_workbook('sites.xlsx')
    logging.info('Workbook loaded successfully.')
except Exception as e:
    logging.error('Failed to find workbook named "sites.xlsx".')
    sys.exit()

# Print sheet names
# console.print(wb.sheetnames) # Used for debugging

# Get sheet
try:
    sheet = wb['Sheet1']
    logging.info('Sheet accessed successfully.')
except Exception as e:
    logging.error('Failed to find sheet named "Sheet1".')
    sys.exit()

# Print records
try:
    for row in track(range(2, sheet.max_row + 1), description='Retrieving lat long data...'):
        address_str = str(sheet['A' + str(row)].value) + ',' + str(sheet['B' + str(row)].value) + ',' + str(sheet['C' + str(row)].value) + ',' + str(sheet['D' + str(row)].value)
        address = get_lat_long(address_str, API_KEY)

        if address[0] is not None:
            sheet['E' + str(row)] = address[0]
            sheet['F' + str(row)] = address[1]
            sheet['G' + str(row)] = address[2]
        else:
            sheet['E' + str(row)] = ''
            sheet['F' + str(row)] = ''
            sheet['G' + str(row)] = ''
            sheet['H' + str(row)] = 'Could not retrieve lat long'

    wb.save('updated_sites.xlsx')
    logging.info('New workbook named "updated_sites.xlsx" saved successfully.')

except Exception as e:
    logging.error(f'Something went wrong: {e}')

